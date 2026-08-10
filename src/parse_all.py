# -*- coding: utf-8 -*-
"""
Step 1: 数据规范化
===============
将所有 Excel 文本块解析为结构化表格。所有数值统一为 0-1 小数格式（精确率、召回率、F1、准确率等）。
百分比格式（如 64.59%）自动转换为小数（0.6459）。

输入: data/ 目录下的所有 Excel 文件
输出: output/data/汇总规范化数据.xlsx + data/汇总规范化数据.xlsx

用法: python scripts/parse_all.py
"""

import re
import openpyxl
import pandas as pd
from pathlib import Path

PROJECT = Path(__file__).resolve().parent.parent
DATA_DIR = PROJECT / "data"
OUTPUT_DIR = PROJECT / "output" / "data"
OUTPUT = OUTPUT_DIR / "汇总规范化数据.xlsx"

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


# ============================================================
# HELPERS
# ============================================================
def extract_metric_ci(text_str):
    """从 '0.94 (0.92–0.96)' 或 '64.59 (55.46–73.11)' 中提取数值与 CI"""
    m = re.search(r'([\d.]+)\s*[\(（]\s*([\d.]+)[–\-]\s*([\d.]+)[\)）]', text_str)
    if m:
        return float(m.group(1)), float(m.group(2)), float(m.group(3))
    return None, None, None


def to_decimal_if_pct(val, lo, hi, col_label=""):
    """如果数值 > 1.01，默认除以 100（排除 RMSE/MAE/SUV/Kappa/Score/Dist 等例外）"""
    if val is None:
        return val, lo, hi
    non_pct_keywords = ["RMSE", "MAE", "SUVmax", "Kappa", "Score", "Dist",
                        "Cohen", "Shot", "AUC-ROC", "AUC-PR", "F1 (Neg)"]
    is_non_pct = any(kw.lower() in str(col_label).lower() for kw in non_pct_keywords)
    if val > 1.01 and not is_non_pct:
        return val / 100, (lo / 100 if lo is not None else None), (hi / 100 if hi is not None else None)
    return val, lo, hi


def convert_cell_to_decimal(cell_value, col_label=""):
    """将单元格值（可能是 '72.8 (71.7–74.0)' 或 '76.0 (64.0–88.0)§' 格式）转为小数"""
    if cell_value is None or not isinstance(cell_value, str):
        return cell_value
    stripped = str(cell_value).strip()
    # 检测尾部标记如 §, †, ‡, *, \n 等
    suffix = ""
    m_suffix = re.search(r'([§†‡*\n]+)$', stripped)
    if m_suffix:
        suffix = m_suffix.group(1).replace('\n', '')
        stripped = stripped[:m_suffix.start()]
    # 匹配: number% (number% – number%) 等变体
    m = re.search(r'^([\d.]+)%?\s*[\(（]\s*([\d.]+)%?\s*[–\-\s]+([\d.]+)%?[\)）]$', stripped)
    if m:
        v = float(m.group(1))
        lo = float(m.group(2))
        hi = float(m.group(3))
        v2, lo2, hi2 = to_decimal_if_pct(v, lo, hi, col_label)
        if v2 != v:
            return f"{v2:.4f} ({lo2:.4f}–{hi2:.4f}){suffix}"
        return f"{v:.2f} ({lo:.2f}–{hi:.2f}){suffix}"
    # Handle standalone number (may have %)
    m = re.search(r'^([\d.]+)%?$', stripped)
    if m:
        v = float(m.group(1))
        v2, _, _ = to_decimal_if_pct(v, None, None, col_label)
        if v2 != v:
            return f"{v2:.4f}{suffix}"
    return cell_value


def convert_string_pct_to_decimal(df):
    """对 DataFrame 中所有 string 类型的百分比单元格进行小数转换（使用位置索引避免重复列名问题）"""
    for col_idx in range(len(df.columns)):
        col_name = str(df.columns[col_idx])
        for idx in range(len(df)):
            val = df.iat[idx, col_idx]
            if isinstance(val, str):
                df.iat[idx, col_idx] = convert_cell_to_decimal(val, col_name)
    return df


# ============================================================
# PARSER 1: PYCM text blocks
# ============================================================
def parse_pycm_block(text, model_name):
    if not text or not isinstance(text, str):
        return []
    results = []
    lines = text.split("\n")
    current_field = None
    n_val = None
    buf = {}

    def ext(line):
        return extract_metric_ci(line)

    for line in lines:
        line = line.strip()
        if line.startswith("字段："):
            current_field = line.replace("字段：", "").strip()
            n_val = None
            buf = {}
        elif current_field and "总样本数" in line:
            m = re.search(r'(\d+)', line)
            n_val = int(m.group(1)) if m else None
        elif current_field and "准确率" in line:
            v, lo, hi = ext(line)
            buf["准确率"], buf["准确率_lo"], buf["准确率_hi"] = v, lo, hi
        elif current_field and "精确率" in line:
            v, lo, hi = ext(line)
            buf["精确率"], buf["精确率_lo"], buf["精确率_hi"] = v, lo, hi
        elif current_field and "召回率" in line:
            v, lo, hi = ext(line)
            buf["召回率"], buf["召回率_lo"], buf["召回率_hi"] = v, lo, hi
        elif current_field and "F1" in line and ("值" in line or "：" in line):
            v, lo, hi = ext(line)
            if v is not None:
                # Convert accuracy if > 1
                acc = buf.get("准确率")
                acc_lo = buf.get("准确率_lo")
                acc_hi = buf.get("准确率_hi")
                if acc is not None and acc > 1.01:
                    acc = acc / 100
                    acc_lo = acc_lo / 100 if acc_lo else None
                    acc_hi = acc_hi / 100 if acc_hi else None

                results.append({
                    "模型": model_name, "字段": current_field, "总样本数": n_val,
                    "准确率": acc, "准确率_lo": acc_lo, "准确率_hi": acc_hi,
                    "精确率": buf.get("精确率"), "精确率_lo": buf.get("精确率_lo"), "精确率_hi": buf.get("精确率_hi"),
                    "召回率": buf.get("召回率"), "召回率_lo": buf.get("召回率_lo"), "召回率_hi": buf.get("召回率_hi"),
                    "F1": v, "F1_lo": lo, "F1_hi": hi,
                })
                current_field = None
                buf = {}
    return results


# ============================================================
# PARSER 2: Baseline / Follow-up
# ============================================================
def parse_baseline_block(text, model_name):
    if not text or not isinstance(text, str):
        return {}
    result = {"模型": model_name}
    m = re.search(r'Weighted F1.*?([\d.]+)\s*[\(（]\s*([\d.]+)[–\-]\s*([\d.]+)[\)）]', text)
    if m:
        result["Weighted_F1"] = float(m.group(1))
        result["Weighted_F1_lo"] = float(m.group(2))
        result["Weighted_F1_hi"] = float(m.group(3))
    m = re.search(r'Accuracy.*?\(%\).*?([\d.]+)\s*[\(（]\s*([\d.]+)[–\-]\s*([\d.]+)[\)）]', text)
    if m:
        result["Accuracy"] = float(m.group(1)) / 100
        result["Accuracy_lo"] = float(m.group(2)) / 100
        result["Accuracy_hi"] = float(m.group(3)) / 100
    m = re.search(r'Per-class F1[^:]*[：:]\s*([\d.]+)\s*/\s*([\d.]+)\s*/\s*([\d.]+)\s*/\s*([\d.]+)', text)
    if m:
        result["F1_Dif"] = float(m.group(1))
        result["F1_Foc"] = float(m.group(2))
        result["F1_Min"] = float(m.group(3))
        result["F1_Mix"] = float(m.group(4))
    return result


def parse_followup_block(text, model_name):
    if not text or not isinstance(text, str):
        return {}
    result = {"模型": model_name}
    m = re.search(r'Accuracy.*?\(%\).*?([\d.]+)\s*[\(（]\s*([\d.]+)[–\-]\s*([\d.]+)[\)）]', text)
    if m:
        result["Accuracy"] = float(m.group(1)) / 100
        result["Accuracy_lo"] = float(m.group(2)) / 100
        result["Accuracy_hi"] = float(m.group(3)) / 100
    m = re.search(r"Cohen's Kappa.*?([\d.]+)\s*[\(（]\s*([\d.]+)[–\-]\s*([\d.]+)[\)）]", text)
    if m:
        result["Cohen_Kappa"] = float(m.group(1))
        result["Cohen_Kappa_lo"] = float(m.group(2))
        result["Cohen_Kappa_hi"] = float(m.group(3))
    m = re.search(r'Per-class F1[^:]*[：:]\s*([\d.]+)\s*/\s*([\d.]+)\s*/\s*([\d.]+)', text)
    if m:
        result["F1_Res"] = float(m.group(1))
        result["F1_Sta"] = float(m.group(2))
        result["F1_Pro"] = float(m.group(3))
    return result


# ============================================================
# PARSER 3: External validation
# ============================================================
def parse_external_block(text):
    if not text or not isinstance(text, str):
        return {}
    results = {}
    for metric, key in [("AUC-ROC", "AUC_ROC"), ("AUC-PR", "AUC_PR"),
                         ("Sensitivity", "Sensitivity"), ("Specificity", "Specificity")]:
        m = re.search(rf'{metric}[^\d]*内部[：:]\s*([\d.]+)\s*[\(（]\s*([\d.]+)[–\-]\s*([\d.]+)[\)）]\s*'
                      rf'外部[：:]\s*([\d.]+)\s*[\(（]\s*([\d.]+)[–\-]\s*([\d.]+)[\)）]\s*P\s*[=＝]\s*([\d.]+|nan)',
                      text, re.IGNORECASE)
        if m:
            # Sensitivity/Specificity are in %
            si = 100 if key in ("Sensitivity", "Specificity") else 1
            results[f"{key}_内部"] = float(m.group(1)) / si
            results[f"{key}_内部_lo"] = float(m.group(2)) / si
            results[f"{key}_内部_hi"] = float(m.group(3)) / si
            results[f"{key}_外部"] = float(m.group(4)) / si
            results[f"{key}_外部_lo"] = float(m.group(5)) / si
            results[f"{key}_外部_hi"] = float(m.group(6)) / si
            try:
                results[f"{key}_P值"] = float(m.group(7))
            except:
                results[f"{key}_P值"] = m.group(7)
    return results


# ============================================================
# MAIN
# ============================================================
def main():
    print("=" * 60)
    print("Step 1: 数据规范化（所有指标转为 0-1 小数格式）")
    print("=" * 60)

    all_pycm = []
    all_baseline = []
    all_followup = []
    all_external = []
    structured_sheets = {}

    for fname in ["临时对应图.xlsx", "临时对应图表.xlsx"]:
        fp = DATA_DIR / fname
        if not fp.exists():
            continue
        print(f"\n--- {fname} ---")
        wb = openpyxl.load_workbook(fp, data_only=True)

        # Sheets 2-4: PYCM text blocks + structured tables
        for s_name in ["Sheet2", "Sheet3", "Sheet4"]:
            if s_name not in wb.sheetnames:
                continue
            ws = wb[s_name]
            model_names = []
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col_idx)
                if cell.value and str(cell.value).strip():
                    model_names.append((col_idx, str(cell.value).strip()))

            for col_idx, model_name in model_names:
                cell = ws.cell(row=2, column=col_idx)
                if cell.value and isinstance(cell.value, str) and len(str(cell.value)) > 500:
                    parsed = parse_pycm_block(str(cell.value), model_name)
                    if parsed:
                        all_pycm.extend(parsed)
                        print(f"  {s_name} {model_name}: {len(parsed)} fields")

            # Structured table (row 5+)
            for r in range(3, min(ws.max_row + 1, 8)):
                cell = ws.cell(row=r, column=1)
                if cell.value and str(cell.value).strip().lower() in ("model",):
                    headers = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
                    data = []
                    for dr in range(r + 1, ws.max_row + 1):
                        row = [ws.cell(row=dr, column=c).value for c in range(1, ws.max_column + 1)]
                        if any(v is not None for v in row):
                            data.append(row)
                    if data:
                        df = pd.DataFrame(data, columns=headers)
                        df = convert_string_pct_to_decimal(df)
                        key = f"模型字段对比_{fname}_{s_name}"
                        structured_sheets[key] = df
                    break

        # Sheet6: Human evaluation
        if "Sheet6" in wb.sheetnames:
            ws = wb["Sheet6"]
            headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            data = []
            for r in range(2, ws.max_row + 1):
                row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
                if any(v is not None for v in row):
                    data.append(row)
            if data:
                df = pd.DataFrame(data, columns=headers)
                df = convert_string_pct_to_decimal(df)
                structured_sheets["人为评估_质量指标"] = df

        # Sheet7: Baseline/Follow-up
        if "Sheet7" in wb.sheetnames:
            ws = wb["Sheet7"]
            for r in range(1, ws.max_row + 1):
                model_cell = ws.cell(row=r, column=1)
                if model_cell.value and str(model_cell.value).strip():
                    model_name = str(model_cell.value).strip()
                    bl = ws.cell(row=r, column=2).value
                    fu = ws.cell(row=r, column=4).value
                    if bl and isinstance(bl, str) and len(str(bl)) > 20:
                        b = parse_baseline_block(str(bl), model_name)
                        if len(b) > 1:
                            already = [x for x in all_baseline if x.get("模型") == model_name]
                            if not already:
                                all_baseline.append(b)
                    if fu and isinstance(fu, str) and len(str(fu)) > 20:
                        f = parse_followup_block(str(fu), model_name)
                        if len(f) > 1:
                            already = [x for x in all_followup if x.get("模型") == model_name]
                            if not already:
                                all_followup.append(f)

        # Sheet8: MRD metrics
        if "Sheet8" in wb.sheetnames:
            ws = wb["Sheet8"]
            headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            data = []
            for r in range(2, ws.max_row + 1):
                row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
                if any(v is not None for v in row):
                    data.append(row)
            if data:
                df = pd.DataFrame(data, columns=headers)
                df = convert_string_pct_to_decimal(df)
                structured_sheets["MRD多模型对比"] = df

        # Sheet9: External validation
        if "Sheet9" in wb.sheetnames:
            ws = wb["Sheet9"]
            for col_idx in [3, 7]:
                cell = ws.cell(row=1, column=col_idx)
                if cell.value and isinstance(cell.value, str) and len(str(cell.value)) > 500:
                    parsed = parse_pycm_block(str(cell.value), f"外部验证_col{col_idx}")
                    if parsed:
                        all_pycm.extend(parsed)
            for sc, suffix in [(1, "set1"), (7, "set7")]:
                h = [ws.cell(row=3, column=c).value for c in range(sc, sc + 4)]
                if h and any(x is not None for x in h):
                    data = []
                    for r in range(4, ws.max_row + 1):
                        row = [ws.cell(row=r, column=c).value for c in range(sc, sc + 4)]
                        if any(v is not None for v in row):
                            data.append(row)
                    if data:
                        df = pd.DataFrame(data, columns=h)
                        df = convert_string_pct_to_decimal(df)
                        structured_sheets[f"外部验证集_Sheet9_{suffix}"] = df

        # Sheet11-13: External validation
        for s_name in ["Sheet11", "Sheet12", "Sheet13"]:
            if s_name not in wb.sheetnames:
                continue
            ws = wb[s_name]
            for col_idx in range(1, 5):
                cell = ws.cell(row=1, column=col_idx)
                if cell.value and isinstance(cell.value, str) and len(str(cell.value)) > 100:
                    if "AUC-ROC" in str(cell.value):
                        ext = parse_external_block(str(cell.value))
                        if ext:
                            all_external.append(ext)

            h = [ws.cell(row=3, column=c).value for c in range(1, ws.max_column + 1)]
            if h and any(x is not None for x in h):
                data = []
                for r in range(4, ws.max_row + 1):
                    row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
                    if any(v is not None for v in row):
                        data.append(row)
                if data:
                    df = pd.DataFrame(data, columns=h)
                    df = convert_string_pct_to_decimal(df)
                    structured_sheets[f"外部验证集_{s_name}"] = df

        # 2-doctor agreement
        if "2医生金标准结果比对" in wb.sheetnames:
            ws = wb["2医生金标准结果比对"]
            # Columns I-L: field name, agreement%, kappa, p-value
            field_names = []
            for r in range(3, ws.max_row + 1):
                f = ws.cell(row=r, column=9)
                if f.value:
                    field_names.append(str(f.value).strip())
            headers = ["字段", "一致率", "Cohen_Kappa", "P_Bowker"]
            data = []
            for r in range(3, 3 + len(field_names)):
                agree = ws.cell(row=r, column=10).value  # %
                kappa = ws.cell(row=r, column=11).value
                pval = ws.cell(row=r, column=12).value
                # Convert agreement from % to decimal
                try:
                    agree_val = float(agree) / 100 if agree else None
                except:
                    agree_val = agree
                try:
                    kappa_val = float(kappa) if kappa else None
                except:
                    kappa_val = kappa
                try:
                    pval_val = float(pval) if pval else None
                except:
                    pval_val = pval
                data.append([field_names[r - 3], agree_val, kappa_val, pval_val])
            if data:
                structured_sheets["2医生金标准结果比对"] = pd.DataFrame(data, columns=headers)

        # Radar chart data
        if "外部验证集雷达图" in wb.sheetnames:
            ws = wb["外部验证集雷达图"]
            headers = ["Metric"] + [str(ws.cell(row=1, column=c).value or f"Col{c}") for c in range(2, ws.max_column + 1)]
            data = []
            for r in range(2, ws.max_row + 1):
                row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
                if any(v is not None for v in row):
                    data.append(row)
            if data:
                df = pd.DataFrame(data, columns=headers)
                df = convert_string_pct_to_decimal(df)
                structured_sheets["外部验证集雷达图数据"] = df

        wb.close()

    # ---- Build data-source mapping ----
    mapping_rows = [
        ["Sheet名称", "来源文件", "原始位置", "对应图表/表格", "说明"],
        ["PYCM字段级指标", "临时对应图.xlsx / 临时对应图表.xlsx", "Sheet2-4 文本块", "Fig3_Heatmap", "22模型×21字段的准确率/精确率/召回率/F1"],
        ["初诊基线评估", "临时对应图表.xlsx", "Sheet7 文本块", "Table 4, Fig1/2 的per-class F1", "9个模型的基线评估（骨髓浸润模式分类）"],
        ["随访评估", "临时对应图表.xlsx", "Sheet7 文本块", "Table 4 follow-up部分", "10个模型的随访评估（缓解/进展/稳定）"],
        ["外部验证集汇总对比", "临时对应图表.xlsx", "Sheet11 文本块", "Table 4 MRD部分", "内部vs外部验证集AUC/Sens/Spec对比"],
        ["人为评估_质量指标", "临时对应图.xlsx", "Sheet6 规范表", "Table 6", "模型输出的工质量评分+幻觉率"],
        ["MRD多模型对比", "临时对应图.xlsx", "Sheet8 规范表", "Table 8, Fig4/5", "8模型MRD评估多维度对比"],
        ["模型字段对比_*", "临时对应图/图表.xlsx", "Sheet2-4 结构化表", "Fig3, Table 6/7/8", "按模型的字段级F1/Accuracy/RMSE对比"],
        ["外部验证集_Sheet9/11/12/13", "临时对应图表.xlsx", "Sheet9-13", "Fig9, Table 9/11", "外部验证集字段级+MRD综合对比"],
        ["2医生金标准结果比对", "临时对应图表.xlsx", "Sheet'2医生金标准结果比对'", "Table 1/2 金标准一致性", "两位放射科医生金标准标注的Kappa一致性"],
        ["外部验证集雷达图数据", "临时对应图表.xlsx", "Sheet'外部验证集雷达图'", "Fig6/Fig9 雷达图", "5指标×3模型的外部验证集数据"],
    ]
    df_mapping = pd.DataFrame(mapping_rows[1:], columns=mapping_rows[0])

    # ---- Post-process: convert all string pct cells in all structured sheets ----
    for name in list(structured_sheets.keys()):
        structured_sheets[name] = convert_string_pct_to_decimal(structured_sheets[name])

    # ---- Write output ----
    print(f"\n{'='*60}")
    print(f"Writing: {OUTPUT}")

    with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
        # Mapping sheet first
        df_mapping.to_excel(writer, sheet_name="数据源映射", index=False)

        if all_pycm:
            df = pd.DataFrame(all_pycm).drop_duplicates()
            df.to_excel(writer, sheet_name="PYCM字段级指标", index=False)
            print(f"  PYCM字段级指标: {len(df)} rows | {df['模型'].nunique()} models | {df['字段'].nunique()} fields")

        if all_baseline:
            pd.DataFrame(all_baseline).to_excel(writer, sheet_name="初诊基线评估", index=False)
            print(f"  初诊基线评估: {len(all_baseline)} models")

        if all_followup:
            pd.DataFrame(all_followup).to_excel(writer, sheet_name="随访评估", index=False)
            print(f"  随访评估: {len(all_followup)} models")

        if all_external:
            pd.DataFrame(all_external).to_excel(writer, sheet_name="外部验证集汇总对比", index=False)
            print(f"  外部验证集汇总对比: {len(all_external)} records")

        for name, df in structured_sheets.items():
            safe_name = name[:31]
            df.to_excel(writer, sheet_name=safe_name, index=False)
            print(f"  {safe_name}: {len(df)} rows")

    # Also copy to data/ directory
    import shutil
    shutil.copy(OUTPUT, DATA_DIR / "汇总规范化数据.xlsx")

    # Export individual CSVs
    if all_pycm:
        pd.DataFrame(all_pycm).drop_duplicates().to_csv(OUTPUT_DIR / "fields_pycm.csv", index=False)
    if all_baseline:
        pd.DataFrame(all_baseline).to_csv(OUTPUT_DIR / "baseline_eval.csv", index=False)
    if all_followup:
        pd.DataFrame(all_followup).to_csv(OUTPUT_DIR / "followup_eval.csv", index=False)
    df_mapping.to_csv(DATA_DIR / "数据源映射表.csv", index=False, encoding="utf-8-sig")

    print(f"\nDone! -> {OUTPUT}")
    print(f"Copy -> {DATA_DIR / '汇总规范化数据.xlsx'}")


if __name__ == "__main__":
    main()
